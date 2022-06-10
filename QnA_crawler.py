import requests
from bs4 import BeautifulSoup
import openpyxl

#엑셀 생성
fpath = r'C:\Users\YS\Desktop\project\crawler\crawling_xlsx\crawling_data.xlsx'
wb = openpyxl.load_workbook(fpath)
ws = wb['한국채식연합QnA'] #wb.create_sheet('한국채식연합QnA')
ws['A1'] = '질문 번호'
ws['B1'] = '질문 제목'
ws['C1'] = '질문 내용'
ws['D1'] = '질문 답변'
ws['E1'] = '작성일'
ws['F1'] = 'url'

rowNum = 2
for pageNum in range(1, 49, 1):
    response = requests.get(f"https://www.vege.or.kr/qna.html?page={pageNum}&page_list=1&&db_name=a_6&kwd=")
    soup = BeautifulSoup(response.content.decode('euc-kr', 'replace'), 'html.parser')
    nums = soup.select('tr.list1 > .list_han_list:first-child')
    dates = soup.select('tr.list1 > .list_han_list:nth-child(4)')
    links = soup.select('tr.list1 > td:nth-child(2) > a')

    for idx, link in enumerate(links):
        ws[f'A{rowNum}'] = nums[idx].text
        ws[f'B{rowNum}'] = link.text
        qlink = 'https://www.vege.or.kr/' + link.attrs['href']
        ws[f'E{rowNum}'] = dates[idx].text
        ws[f'F{rowNum}'] = qlink

        qrequest = requests.get(qlink)
        qsoup = BeautifulSoup(qrequest.content.decode('euc-kr', 'replace'), 'html.parser')
        question = qsoup.select_one('.list_han3')
        answer = qsoup.select('font.list_han')

        ws[f'C{rowNum}'] = question.text if question is not None else ""
        if(len(answer) > 1):
            ws[f'D{rowNum}'] = answer[2].text
        else:
            ws[f'D{rowNum}'] = ''

        rowNum += 1

#엑셀 저장
wb.save(fpath)